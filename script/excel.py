# pip install openpyxl

import datetime
import os
import openpyxl

from openpyxl import load_workbook

def walkFile(target_path, key):
    key = key.encode()
    total_num = 0

    for base_path, folder_list, file_list in os.walk(target_path):
        for file_name in file_list:
            file_path = os.path.join(base_path, file_name)
            file_ext = file_path.rsplit(".", maxsplit=1)
            if len(file_ext) != 2:
                # 
                continue

            if file_ext[1] != "java" and file_ext[1] != "xml":
                # 
                continue

            file_num = 0
            with open(file_path, "rb") as f:
                lines = f.readlines()
                for line in lines:
                    if key in line:
                        file_num += 1
            total_num += file_num

    return total_num


print("start")

wb = load_workbook("C:\\Users\\admin\\Desktop\\test.xlsx")
sheet = wb.worksheets[3]

for clo in range(29, 148):
    batchName = sheet.cell(5, clo).value
    filePathName = sheet.cell(6, clo).value
    if batchName != None and filePathName != None:
        print(batchName + ":" + filePathName)
        for row in range(7, 218):
            tableName = sheet.cell(row, 6).value
            targetCell = sheet.cell(row, clo)
            #
            num = walkFile(filePathName, tableName)
            targetCell.value = "" if num > 0 else ""
            #

wb.save(filename="C:\\Users\\admin\\Desktop\\test.xlsx")

print("end")
